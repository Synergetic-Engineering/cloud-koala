import decimal
import os
import uuid

from openpyxl import load_workbook


DEFAULT_CONFIG_SHEET_NAME = 'cloud-koala-config'


def _find_input_output(workbook):
    input_datalists = {}
    output_datalists = {}
    for sheet in workbook:
        sheet_title = sheet.title.encode('utf-8')
        input_datalists[sheet_title] = []
        output_datalists[sheet_title] = []

        for col in sheet.iter_cols():
            for cell in col:
                if cell.value is None:
                    continue
                elif isinstance(cell.value, unicode):
                    continue
                elif isinstance(cell.value, (float, long)):
                    input_datalists[sheet_title].append(str(cell.column) + str(cell.row))
                elif isinstance(cell.value, str) and cell.value.startswith('='):
                    output_datalists[sheet_title].append(str(cell.column) + str(cell.row))

        # If sheet is empty, delete it from dictionary
        if input_datalists[sheet_title] == []:
            del input_datalists[sheet_title]
        if output_datalists[sheet_title] == []:
            del output_datalists[sheet_title]
    return input_datalists, output_datalists


def _generate_config_rows(cell_type, datalists):
    for elements in datalists:
        for coords in datalists[elements]:
            yield cell_type, elements, coords


def create_config_sheet(file_string, config_sheet_name=None):
    # XXX HACK this was required for koala but haven't worked out if it's required for openpyxl yte
    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    dummy_excel_file_name = '/tmp/temp_excel_file_{}.xlsx'.format(uuid.uuid4().hex)
    with open(dummy_excel_file_name, 'wb') as f:
        f.write(file_string)

    if config_sheet_name is None:
        config_sheet_name = DEFAULT_CONFIG_SHEET_NAME

    workbook = load_workbook((dummy_excel_file_name), data_only=False)
    input_datalists, output_datalists = _find_input_output(workbook)
    rows = [('type', 'sheet', 'cell')]
    rows.extend(_generate_config_rows('input', input_datalists))
    rows.extend(_generate_config_rows('output', output_datalists))
    # XXX should check if config_sheet_name already exists and maybe fail if it does
    config_sheet = workbook.create_sheet(config_sheet_name)
    for row in rows:
        config_sheet.append(row)
    workbook.save(dummy_excel_file_name)
    with open(dummy_excel_file_name, 'rb') as f:
        new_file_string = f.read()

    # cleans up previous workaround
    os.remove(dummy_excel_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')

    return new_file_string


def get_config_info(file_string, config_sheet_name=None, has_header_row=True):
    # XXX HACK this was required for koala but haven't worked out if it's required for openpyxl yte
    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    dummy_excel_file_name = '/tmp/temp_excel_file_{}.xlsx'.format(uuid.uuid4().hex)
    with open(dummy_excel_file_name, 'wb') as f:
        f.write(file_string)

    if config_sheet_name is None:
        config_sheet_name = DEFAULT_CONFIG_SHEET_NAME

    # load with openpyxl and read the `config_sheet_name` sheet
    workbook = load_workbook((dummy_excel_file_name), data_only=False)
    if config_sheet_name not in workbook:
        return {'err': 'Config sheet `{}` not found'.format(config_sheet_name)}
    config_sheet = workbook[config_sheet_name]

    config_info = []
    for i, row in enumerate(config_sheet.iter_rows()):
        if i == 0:
            if has_header_row:
                headers = [cell.value for cell in row]
                continue
            else:
                headers = ['header{}'.format(i+1) for i in range(len(row))]
        dynamodb_type_safe_row = []
        for cell in row:
            cell_value = cell.value
            if isinstance(cell_value, float):
                cell_value = decimal.Decimal(cell_value)
            dynamodb_type_safe_row.append(cell_value)
        config_info.append(dict(zip(headers, dynamodb_type_safe_row)))
    print 'config_info', config_info

    # cleans up previous workaround
    os.remove(dummy_excel_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')

    return config_info
