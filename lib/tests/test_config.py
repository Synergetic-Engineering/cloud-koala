import os

from openpyxl import load_workbook

import lib.config
from lib.tests import util


def test_create_config_sheet():
    file_string = util.get_file_string()

    new_file_string = lib.config.create_config_sheet(file_string)

    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    dummy_excel_file_name = '/tmp/temp_excel_file.xlsx'
    with open(dummy_excel_file_name, 'wb') as f:
        f.write(new_file_string)

    workbook = load_workbook((dummy_excel_file_name), data_only=False)
    config_sheet = workbook['cloud-koala-config']

    rows = []
    for row in config_sheet.iter_rows():
        for cell in row:
            rows.append(cell.value.encode('utf-8'))

    expecting = ['type', 'sheet', 'cell',
                 'input', 'Sheet1', 'B2',
                 'input', 'Sheet1', 'B3',
                 'input', 'Sheet1', 'B4',
                 'input', 'Sheet1', 'B5',
                 'input', 'Sheet1', 'B6',
                 'input', 'Sheet1', 'B7',
                 'input', 'Sheet1', 'B8',
                 'input', 'Sheet1', 'B9',
                 'input', 'Sheet1', 'B10',
                 'input', 'Sheet1', 'B11',
                 'output', 'Sheet1', 'B12']
    assert rows == expecting

    os.remove(dummy_excel_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')


def test_get_config_info():
    file_string = util.get_file_string(file_name='lib/tests/test_with_existing_config_sheet.xlsx')

    config_info = lib.config.get_config_info(file_string, 'config_sheet_test')

    expected_headers = [
        'type',
        'variable_name',
        'sheet',
        'table',
        'cell',
        'units',
        'default_value',
        'description',
        ]
    expected_rows = [
        ('input', 'i1', 'Sheet1', 'test table', 'B2', 'm', 1, 'first value'),
        ('input', 'i2', 'Sheet1', 'test table', 'B3', 'm', 2, 'second value'),
        ('output', 'avg', 'Sheet1', 'test table', 'B12', 'm', 5.5, 'average'),
    ]

    assert config_info == [dict(zip(expected_headers, row)) for row in expected_rows]
