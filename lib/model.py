import os
import time
import uuid

import boto3
import botocore

from openpyxl import load_workbook

from koala.ExcelCompiler import ExcelCompiler
from koala.Spreadsheet import Spreadsheet
from koala.ExcelError import ExcelError


# XXX is having these up at the module level like this best practice?
# - also worth considering how we will mock them for testing
dynamodb = boto3.resource('dynamodb')
table = dynamodb.Table(os.environ['DYNAMODB_TABLE'])
s3 = boto3.resource('s3')
bucket = s3.Bucket(os.environ['S3_BUCKET'])

def _update_bucket_parameter(model_id, param_name, content):
    table.update_item(
        Key={'model_id': model_id},
        UpdateExpression='SET #name = :value',
        ExpressionAttributeNames={"#name":param_name},
        ExpressionAttributeValues={":value":content}
    )

def _move_bucket_object(start_folder, end_folder, model_id):
    status = "200"
    try:
        bucket.put_object(
            Body="",
            Key='%s/%s'%(end_folder, model_id), ## big changes
        )
        obj = bucket.Object('%s/%s'%(end_folder, model_id))
        obj.copy({
            'Bucket': os.environ['S3_BUCKET'],
            'Key': '%s/%s'%(start_folder, model_id)
        })
    except botocore.exceptions.ClientError as err:
        if err.response['Error']['Code'] == "404":
            status = "404"
    #TODO Add Exception Handling for specific server not found
    # except Exception as err:
    #     if err:
    #         status = "404"
        else:
            status = "Error"
    finally:
        bucket.delete_objects(Delete={
            'Objects': [{'Key': '%s/%s'%(start_folder, model_id)}]
        })
    return status

def get_model_config(file_string):
    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    temp_id = uuid.uuid4().hex
    dummy_excel_file_name = '/tmp/temp_excel_file_{}.xlsx'.format(temp_id)
    with open(dummy_excel_file_name, 'wb') as fp:
        fp.write(file_string)
    # TODO (eventually integrate with add_or_update_model and maybe make
    # this a helper function -> _get_model_config)

    def find_input_output():
        for ws in wb:
            input_datalists[ws.title.encode('utf-8')] = []
            output_datalists[ws.title.encode('utf-8')] = []

            for col in ws.iter_cols():
                for cell in col:
                    if cell.value is None:
                        continue
                    elif type(cell.value) == unicode:
                        continue
                    elif (type(cell.value) == float) or (type(cell.value) == long):
                        input_datalists[ws.title.encode('utf-8')].append(str(cell.column) + str(cell.row))
                    elif (type(cell.value) == str) and ((cell.value).startswith('=')):
                        output_datalists[ws.title.encode('utf-8')].append(str(cell.column) + str(cell.row))

            # If sheet is empty, delete it from dictionary
            if input_datalists[ws.title.encode('utf-8')] == []:
                del input_datalists[ws.title.encode('utf-8')]
            if output_datalists[ws.title.encode('utf-8')] == []:
                del output_datalists[ws.title.encode('utf-8')]
        return input_datalists
        return output_datalists

    def write_sheet_input():
        label_array_in = input_datalists.keys()
        for j, elements in enumerate(label_array_in):
            for i, coords in enumerate(input_datalists[elements]):
                rows.append(('input', input_datalists.keys()[j], input_datalists[elements][i]))
        return rows

    def write_sheet_output():
        label_array_out = output_datalists.keys()
        for j, elements in enumerate(label_array_out):
            for i, coords in enumerate(output_datalists[elements]):
                rows.append(('output', output_datalists.keys()[j], output_datalists[elements][i]))
        return rows

    filename = dummy_excel_file_name
    wb = load_workbook((filename), data_only=False)
    input_datalists = {}
    output_datalists = {}
    find_input_output()
    rows = [('type', 'sheet', 'cell')]
    write_sheet_input()
    write_sheet_output()
    config_sheet = wb.create_sheet("cloud-koala-config")

    for row in rows:
        config_sheet.append(row)

    wb.save(dummy_excel_file_name)
    with open(dummy_excel_file_name, 'rb') as fp:
        new_file_string = fp.read()

    os.remove(dummy_excel_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')
    return new_file_string


def list_models():
    """
    List the available models in S3
    """
    result = table.scan()
    return {'models': result['Items']}


def add_or_update_model(file_name, file_string, model_id=None):
    """
    Serialise the xlsx_file and upload to S3, return the ID
    """
    key = {'model_id': model_id or uuid.uuid4().hex}
    if model_id is not None:
        item_record = table.get_item(Key=key).get('Item', {})
    if model_id is None or not item_record:
        item_record = {
            'model_id': key['model_id'],
            'created_at': str(int(time.time())),
            'version': str(0),
            'compilation_status':'Waiting',
        }
    item_record['file_name'] = file_name
    item_record['version'] = str(int(item_record['version'])+1)
    item_record['updated_at'] = str(int(time.time()))
    table.put_item(Item=item_record)
    bucket.put_object(Key='excel_uploads/{}'.format(item_record['model_id']), Body=file_string)
    return key['model_id']


def get_model(model_id):
    """
    Get information about the serialised model and Excel file
    e.g. inputs, outputs
    maybe allow for optional level of detail to this?
    """
    result = table.get_item(Key={'model_id': model_id})
    return result.get('Item')


def delete_model(model_id):
    """
    Delete the model record from S3 and either archive or delete related files in S3
    """
    table.delete_item(Key={'model_id': model_id})
    status = _move_bucket_object('excel_uploads', 'excel_uploads_archive', model_id)
    if _move_bucket_object('compiled_models', 'compiled_models_archive', model_id) == "Error":
        return {'status': "Error"}
    return {'status': status}


def run_model(model_id, input_dict, output_names):
    """
    Load the model, set the inputs and return the calculated outputs
    TODO get this working roughly following these steps
      - Get serialised model from S3
      - Load model with koala
      - Extract inputs from payload
      - Set the inputs in the model
      - Extract required outputs from payload (all outputs if none specifically requested)
      - Get the required outputs from the model
      - Build and return response
    """
    # see if compiled model compiled
    try:
        compliled_string = bucket.Object('compiled_models/{}'.format(model_id)).get()['Body'].read()
    except botocore.exceptions.ClientError as err:
        return err.response['Error']['Code']
    # XXX HACK Workaround needed for koala spreadsheet loading API
    # - need to write the file to a temp location for koala to read it...
    # - FIX = koala.Spreadsheet / koala.serialize should be updated to take the file contents in directly
    if not os.path.exists('/tmp'):
        os.mkdir('/tmp')
    dummy_file_name = '/tmp/temp_{}.gzip'.format(model_id)
    with open(dummy_file_name, 'wb') as fp:
        fp.write(compliled_string)
    sp = Spreadsheet.load(dummy_file_name)
    for name, value in input_dict.iteritems():
        sp.set_value(name, value)
    results = {}
    for name in output_names:
        results[name] = sp.evaluate(name)
    # Cleanup previous workaround
    os.remove(dummy_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')
    return results


def compile_model(model_id):
    compliled_string = bucket.Object('excel_uploads/{}'.format(model_id)).get()['Body'].read()
    _update_bucket_parameter(model_id, "compilation_status", "Compiling")
    # XXX HACK Workaround needed for koala spreadsheet loading API
    # - need to write the file to a temp location for koala to read it...
    # - then need to write koala compiled file from a temp location...
    # - FIX = koala.Spreadsheet / koala.serialize should be updated to take the file contents in directly
    return_str = ''
    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    dummy_excel_file_name = '/tmp/temp_excel_file_{}.xlsx'.format(model_id)
    with open(dummy_excel_file_name, 'wb') as fp:
        fp.write(compliled_string)
    try:
        compiler = ExcelCompiler(dummy_excel_file_name)
        sp = compiler.gen_graph()
    except IOError as err:
        print err
        _update_bucket_parameter(model_id, "compilation_status", "Failed (Invalid Excel File)")
        return_str = 'model {} did not compile'.format(model_id)
    except Exception as err:
        if str(err)=='File is not a zip file':
            _update_bucket_parameter(model_id, "compilation_status", "Failed (Invalid File Type)")
            print err
        else:
            _update_bucket_parameter(model_id, "compilation_status", "Failed (Generic Error)")
            print err
        return_str = 'model {} did not compile'.format(model_id)
    else:
        dummy_compiled_file_name = '/tmp/temp_compiled_file_{}.gzip'.format(model_id)
        sp.dump(dummy_compiled_file_name)
        with open(dummy_compiled_file_name, 'rb') as fp:
            compiled_file_string = fp.read()
        # Write compiled file to S3 and update dynamodb record
        bucket.put_object(Key='compiled_models/{}'.format(model_id), Body=compiled_file_string)
        _update_bucket_parameter(model_id, "compilation_status", "Compiled")
        # Cleanup previous workaround
        os.remove(dummy_compiled_file_name)
        return_str = 'model {} compiled'.format(model_id)
    finally:
        # Cleanup previous workaround
        os.remove(dummy_excel_file_name)
        if not os.listdir('/tmp'):
            os.rmdir('/tmp')
        return return_str
