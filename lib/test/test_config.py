import os

from openpyxl import load_workbook

# XXX this needs to happen before importing other stuff to set
# env variables
from lib.test import util

import boto3
boto3.setup_default_session(aws_access_key_id='123', aws_secret_access_key='123', region_name='ap-southeast-2')

import lib.config


@util.setup_mock_resources
def test_create_config_sheet():
    with open('lib/test/test.xlsx', 'rb') as f:
        file_string = f.read()

    new_file_string = lib.config.create_config_sheet(file_string)

    if not os.path.exists('/tmp'): # mainly required for dev / test environment
        os.mkdir('/tmp')
    dummy_excel_file_name = '/tmp/temp_excel_file.xlsx'
    with open(dummy_excel_file_name, 'wb') as fp:
        fp.write(new_file_string)

    workbook = load_workbook((dummy_excel_file_name), data_only=False)
    config_sheet = workbook['cloud-koala-config']

    rows = []
    for row in config_sheet.iter_rows():
        for cell in row:
            rows.append(cell.value.encode('utf-8'))

    expecting = ['type', 'sheet', 'cell',
                 'input', 'Sheet1', 'B2',
                 'input', 'Sheet1', 'B3',
                 'input', 'Sheet1', 'B4',
                 'input', 'Sheet1', 'B5',
                 'input', 'Sheet1', 'B6',
                 'input', 'Sheet1', 'B7',
                 'input', 'Sheet1', 'B8',
                 'input', 'Sheet1', 'B9',
                 'input', 'Sheet1', 'B10',
                 'input', 'Sheet1', 'B11',
                 'output', 'Sheet1', 'B12']
    assert rows == expecting

    os.remove(dummy_excel_file_name)
    if not os.listdir('/tmp'):
        os.rmdir('/tmp')
